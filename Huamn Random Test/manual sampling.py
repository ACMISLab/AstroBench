import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path

csv_files = [
    r"your_csv",
]

output_xlsx = r"your_path.xlsx"

SAMPLE_N = 5
INDEX_COL = "Index"
IMAGE_LABEL_COL = "Image Lable"


df_first = pd.read_csv(csv_files[0])

if IMAGE_LABEL_COL in df_first.columns:
    df_candidate = df_first[df_first[IMAGE_LABEL_COL] == "No"]
    print("The Image Label column was detected; sampling will only be performed from rows where Image Label == 'No'.")
else:
    df_candidate = df_first
    print("No Image Label column was detected; by default, all data is available for sampling.")

if len(df_candidate) < SAMPLE_N:
    raise ValueError(
        f"Only {len(df_candidate)} data points are available for sampling, which is less than {SAMPLE_N}."
    )

sample_df = df_candidate.sample(n=SAMPLE_N, random_state=None)

selected_index_values = sample_df[INDEX_COL].tolist()

print("Final sampling index:", selected_index_values)

merged_dfs = []

for csv_path in csv_files:
    df = pd.read_csv(csv_path)

    aligned_df = df[df[INDEX_COL].isin(selected_index_values)]

    aligned_df = (
        aligned_df
        .set_index(INDEX_COL)
        .loc[selected_index_values]
        .reset_index()
    )
    aligned_df["source_csv"] = Path(csv_path).name

    merged_dfs.append(aligned_df)

final_df = pd.concat(merged_dfs, ignore_index=True)

double_rows = final_df.loc[final_df.index.repeat(2)].reset_index(drop=True)

double_rows.to_excel(output_xlsx, index=False)

wb = load_workbook(output_xlsx)
ws = wb.active

max_row = ws.max_row
max_col = ws.max_column

row = 2
while row <= max_row:
    for col in range(1, max_col + 1):
        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row + 1,
            end_column=col
        )
        ws.cell(row=row, column=col).alignment = Alignment(
            vertical="center",
            horizontal="center"
        )
    row += 2

wb.save(output_xlsx)

print("Output path:", output_xlsx)
